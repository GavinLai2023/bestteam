from __future__ import annotations

import codecs
import csv
import io
import re
import xml.etree.ElementTree as ET
import zipfile
from pathlib import Path
from typing import List, Optional
from xml.sax.saxutils import escape, quoteattr

from ..exceptions import ConfigurationError

_TEXT_SUFFIXES = {".txt", ".md", ".csv", ".json", ".yaml", ".yml", ".log"}

SUPPORTED_SUFFIXES = frozenset(
    {".pdf", ".xlsx", ".xlsm", ".docx", ".xml"} | _TEXT_SUFFIXES
)


def parse_file(path: str) -> str:
    """Extract text content from a file.

    Supports PDF (text extraction), Excel (.xlsx/.xlsm, rendered as CSV rows),
    Word (.docx, including tables), XML (structural rendering of tags,
    attributes, and text, including mixed-content tail text and namespace
    prefixes -- comments and processing instructions are dropped, matching
    ElementTree's default parser), and common plain-text formats (.txt, .md,
    .csv, .json, .yaml). Legacy .xls (BIFF) is not supported -- the openpyxl
    backend reads only the modern Office Open XML formats.

    This tool reads whatever local path it is given, with no sandboxing —
    the same trust boundary as `http_get` fetching arbitrary URLs. If this
    tool is exposed to an LLM agent, the caller is responsible for
    constraining which paths the agent can be prompted to access.

    Args:
        path: Absolute or relative path to the file to parse.

    Returns:
        Extracted text content as a single string.
    """
    file_path = Path(path)
    if not file_path.exists():
        raise ConfigurationError(f"File not found: {path}")

    # `lenient_text` deliberately left at its strict default: a mis-encoded
    # file here is an operator's own document, and a knowledge base is better
    # served by skipping it with a warning than by silently ingesting mojibake
    # chunks nobody can search. Attachments make the opposite trade -- see
    # `parse_bytes`.
    return parse_bytes(file_path.read_bytes(), file_path.name)


def parse_bytes(data: bytes, filename: str, *, lenient_text: bool = False) -> str:
    """Extract text from a file's bytes, dispatching on `filename`'s suffix.

    The byte-based entry point exists because email attachments are named by
    whoever sent the message. `parse_file` reads whatever path it is given,
    with no sandboxing -- a contract that is fine for knowledge bases, where an
    operator chooses the paths, and exactly wrong for mail. Parsing from bytes
    means there is no path for an attacker-chosen name to become.

    `filename` is used for dispatch and for the rendered header only; it is
    never resolved, opened, or joined to anything.

    Args:
        data: The file's raw bytes.
        filename: The file's name, used only for suffix dispatch and headers.
        lenient_text: Replace undecodable bytes in a plain-text file instead of
            raising. Off by default, so this entry point and `parse_file` agree
            and a mis-encoded document is reported rather than silently
            mangled. The email attachment path turns it on: a sender can name
            anything `.txt`, and one bad attachment must not fail a customer's
            whole run. Affects only plain text -- the binary parsers raise
            either way.

    Returns:
        Extracted text content as a single string.
    """
    suffix = Path(filename).suffix.lower()

    if suffix == ".pdf":
        return _parse_pdf_bytes(data, filename)
    if suffix in (".xlsx", ".xlsm"):
        return _parse_excel_bytes(data, filename)
    if suffix == ".docx":
        return _parse_docx_bytes(data, filename)
    if suffix == ".xml":
        return _parse_xml_bytes(data, filename)
    if suffix == ".csv":
        return _parse_csv_bytes(data, filename, lenient_text=lenient_text)
    if suffix in _TEXT_SUFFIXES:
        return _decode_text(data, filename, lenient=lenient_text)

    raise ConfigurationError(
        f"Unsupported file type '{suffix}'. "
        f"Supported types: .pdf, .xlsx, .xlsm, .docx, .xml, {', '.join(sorted(_TEXT_SUFFIXES))}"
    )


def _decode_text(data: bytes, name: str, *, lenient: bool = False) -> str:
    """Decode plain-text bytes, trying the encodings customers actually send.

    UTF-8 first, then the BOM-announced encodings, then GB18030 -- because a
    plain-text document produced on a Chinese Windows machine usually is not
    UTF-8. Notepad and Excel's "CSV (comma delimited)" export both write GBK;
    Excel's "Unicode text" export writes UTF-16 with a BOM. Each of those used
    to reach the customer as a raw `UnicodeDecodeError` naming a byte offset.

    Still strict about the thing strictness was protecting: a document that
    decodes as none of these is reported, not stored as mojibake. GB18030 is
    the reason the CJK check exists -- it is a very permissive codec, and any
    Western-encoded document whose high bytes happen to form valid pairs would
    otherwise stop failing loudly and start being indexed as Chinese nonsense.
    A GB18030 result is therefore accepted only when it contains a *run* of
    CJK characters, which is the only reason we try that codec at all. The run
    is what makes the check mean anything: Latin-1 Spanish or Swedish spells
    each accented letter with one high byte, which pairs with the ASCII byte
    after it into a single lone Han character, so "contains CJK" is satisfied
    by ordinary Western prose. Chinese text comes in runs; that accident does
    not. The cost is that a genuinely GB18030-encoded document with no two
    adjacent Chinese characters is refused -- a loud refusal the customer
    fixes by re-saving as UTF-8, chosen over a silent one nobody would notice.

    `lenient` keeps the attachment path's contract: a sender can name anything
    `.txt`, and a decode error escaping into the poller would fail a whole run
    over one bad attachment. It is the last resort, though, not the first --
    an attachment that IS GB18030 arrives as its own text, not as U+FFFDs.

    Line endings are translated either way, because the path-based reader this
    replaces opened files in text mode (universal newlines) and its output must
    not change.
    """
    decoded = _decoded_or_none(data)
    if decoded is None:
        if not lenient:
            raise ConfigurationError(
                f"Could not read '{name}': it is not UTF-8, GB18030 or UTF-16 text. "
                "Open it and save it again as UTF-8, then upload it once more."
            )
        decoded = data.decode("utf-8", errors="replace")
    return decoded.replace("\r\n", "\n").replace("\r", "\n")


# Two or more adjacent CJK Unified Ideographs (Extension A included) -- what
# a GB18030 document is for. See `_decode_text` for why a decode that yields
# no such run is rejected, and why a lone one is not evidence of anything.
_CJK_RUN_RE = re.compile(r"[\u3400-\u4dbf\u4e00-\u9fff]{2,}")


def _decoded_or_none(data: bytes) -> Optional[str]:
    """The first encoding in the chain that reads `data`, or None.

    A BOM is an explicit declaration by whoever wrote the file, so it is
    honoured before anything is guessed. It is also consumed: a UTF-8 BOM is
    read with `utf-8-sig` rather than `utf-8`, which would otherwise leave it
    as the document's first character, where it becomes part of whatever the
    first line is -- a heading, or a spreadsheet's first column name.

    A BOM is a claim, not a guarantee, so a file that opens with one and then
    fails to decode still falls through to the rest of the chain. The order
    the BOMs are tested in matters, though: one is a prefix of another.
    """
    if data.startswith(codecs.BOM_UTF8):
        candidates = ("utf-8-sig", "gb18030")
    elif data.startswith((codecs.BOM_UTF32_LE, codecs.BOM_UTF32_BE)):
        # Ahead of UTF-16, because a UTF-32 LE BOM *begins with* a UTF-16 LE
        # BOM -- and a UTF-32 file read as UTF-16 does not fail, it decodes
        # into the same text with a NUL between every character.
        candidates = ("utf-32", "utf-8", "gb18030")
    elif data.startswith((codecs.BOM_UTF16_LE, codecs.BOM_UTF16_BE)):
        candidates = ("utf-16", "utf-8", "gb18030")
    else:
        candidates = ("utf-8", "gb18030")
    for encoding in candidates:
        try:
            decoded = data.decode(encoding)
        except (UnicodeDecodeError, ValueError):
            continue
        if encoding == "gb18030" and not _CJK_RUN_RE.search(decoded):
            continue
        return decoded
    return None


# The delimiter `_parse_pdf_bytes` joins a PDF's pages with, so a page boundary
# survives into the extracted text. Public and owned by the producer, because
# the consumer (`core/knowledge_base.py`, which chunks a PDF per page) has to
# split on exactly what was written here.
PAGE_BREAK = "\f"


def _parse_pdf_bytes(data: bytes, name: str) -> str:
    try:
        import pypdf
    except ImportError as exc:
        raise ConfigurationError(
            "Parsing PDF files requires the 'pypdf' package. "
            "Install it with: pip install 'bestteam[tools-files]'"
        ) from exc

    reader = pypdf.PdfReader(io.BytesIO(data))
    # Pages are joined with a form feed rather than a blank line, so the page
    # boundary survives into the extracted text and a knowledge base can chunk
    # per page and cite an exact `p.N` (see `core/knowledge_base.py`'s
    # `_chunk_document`). That makes the form feed a delimiter, so a page whose
    # own extracted text contains one has to give it up -- otherwise it reads
    # as a page break and silently shifts the number of every page after it,
    # which is the one way the "exact p.N" guarantee could be quietly false.
    # Replaced with a space, not dropped, so words either side stay separated.
    # The email attachment path reads the same string; a form feed is
    # whitespace, so nothing there changes but the separator a model sees
    # between two pages.
    pages = [(page.extract_text() or "").replace(PAGE_BREAK, " ") for page in reader.pages]
    header = f"[PDF: {name} — {len(pages)} page(s)]\n"
    return header + PAGE_BREAK.join(pages)


def _parse_docx_bytes(data: bytes, name: str) -> str:
    try:
        import docx
        from docx.table import Table
    except ImportError as exc:
        raise ConfigurationError(
            "Parsing Word files requires the 'python-docx' package. "
            "Install it with: pip install 'bestteam[tools-files]'"
        ) from exc

    document = docx.Document(io.BytesIO(data))

    # Body order, not "all paragraphs then all tables". A table appended after
    # the prose it belongs to reads as if it were an appendix: the rows survive
    # but the sentence introducing them is somewhere else entirely, and the
    # knowledge base chunks it that way too. `iter_inner_content` yields
    # paragraphs and tables interleaved as the document has them (python-docx
    # >= 1.1, which is why `pyproject.toml` floors it there).
    parts: List[str] = []
    paragraph_run: List[str] = []
    table_number = 0

    def _flush_paragraphs() -> None:
        if paragraph_run:
            parts.append("\n".join(paragraph_run))
            paragraph_run.clear()

    for item in document.iter_inner_content():
        if isinstance(item, Table):
            _flush_paragraphs()
            table_number += 1
            rows = [
                ",".join(_one_line(cell.text) for cell in row.cells)
                for row in item.rows
            ]
            # A row with no text in any cell is dropped rather than rendered.
            # It carries nothing a search could match, and in a ONE-column
            # table it renders as the empty string -- which is the blank line
            # `core/knowledge_base.py::_docx_segments` reads as the end of the
            # block, so a spacer row would drop every row after it out of the
            # table and index it as prose with no header and no citation.
            # Dropping them is what makes that terminator unambiguous.
            parts.append(
                f"[Table {table_number}]\n"
                + "\n".join(row for row in rows if row.replace(",", "").strip())
            )
        elif item.text.strip():
            paragraph_run.append(_docx_paragraph_line(item))

    _flush_paragraphs()
    return f"[Word: {name}]\n" + "\n\n".join(parts)


# The heading shape written here and read back out in
# `core/knowledge_base.py` (`_MARKDOWN_HEADING_RE` is this object). Public and
# owned by the producer for the same reason `PAGE_BREAK` is: the consumer has
# to match exactly what was written, and a copy on each side could drift --
# here the drift would be silent, because an unescaped collision only shows up
# as a wrong citation.
MARKDOWN_HEADING_RE = re.compile(r"^#{1,4} +(.+?)\s*$", re.M)

# The bracketed line every tabular parser writes ahead of a block of CSV-style
# rows -- one per Excel sheet, one per Word table, one for a whole CSV file.
# Public and owned by the producer for the same reason as the two above: the
# consumer splits on exactly this, and the writer has to escape exactly what
# the reader matches.
TABLE_MARKER_RE = re.compile(r"^\[(Sheet: [^\]\n]*|CSV: [^\]\n]*|Table \d+)\]$", re.M)


def _escape_marker_shaped(text: str) -> str:
    """Backslash-escape any line that would be read back as a table marker.

    The heading escape's twin, and the same defect underneath: rendering a
    document's structure as a shape makes that shape meaningful, so a line the
    document's own content happens to spell the same way is ambiguous. A Word
    paragraph reading `[Table 2]`, or a one-column row reading `[CSV: x.csv]`,
    opened a block that swallowed the prose or rows after it -- and cited them
    as a table, or as a document the collection does not contain.

    Word's table rows need none, and do not get any: `_docx_segments` runs a
    block to the blank line after it and ignores a marker found inside one, so
    a cell reading `[Table 2]` is already just a cell. A sheet or a CSV is
    split marker to marker, which is why their rows do need it.
    """
    return TABLE_MARKER_RE.sub(lambda match: "\\" + match.group(0), text)


# Word heading styles are named `Heading 1` … `Heading 9` (plus `Title` for the
# document title). Only the first four map to a Markdown level the knowledge
# base's separators know about (`_MARKDOWN_SEPARATORS` stops at `####`), so
# anything deeper clamps rather than emitting a `#####` the chunker would treat
# as ordinary prose.
_MAX_MARKDOWN_HEADING_LEVEL = 4
_DOCX_HEADING_RE = re.compile(r"^Heading (\d+)$")


def _docx_paragraph_line(paragraph) -> str:
    """Render one Word paragraph, promoting its heading style to Markdown.

    A `.docx` carries its structure in paragraph *styles*, which the old
    text-only extraction dropped -- so a Word document, unlike a Markdown one,
    could never cite the section a chunk came from. Emitting `#` lines lets the
    existing `_headings_for` machinery in `core/knowledge_base.py` apply
    unchanged, and is also the shape a heavier parser (docling's
    `export_to_markdown`) would produce.

    A style *named* like a heading is trusted, including a custom one: the name
    is what the author chose to call it. Ordinary prose keeps its text
    unchanged apart from the two escapes below -- only a heading is normalised
    to one line, because a `#` line broken in two would leave the second half
    as an orphan paragraph.
    """
    style = getattr(paragraph.style, "name", None) or ""
    if style == "Title":
        level = 1
    else:
        match = _DOCX_HEADING_RE.match(style)
        if not match:
            return _escape_marker_shaped(_escape_heading_shaped(paragraph.text))
        level = min(max(int(match.group(1)), 1), _MAX_MARKDOWN_HEADING_LEVEL)
    return f"{'#' * level} {_one_line(paragraph.text)}"


def _escape_heading_shaped(text: str) -> str:
    """Backslash-escape any line of ordinary prose that would be read back as a
    generated heading.

    Rendering Word's heading styles as `#` lines makes the shape meaningful, so
    a Normal-styled paragraph that happens to begin `# ` becomes ambiguous:
    `core/knowledge_base.py` would cite it as the section its chunk opens under
    and cut a chunk boundary at it. The escape is Markdown's own, applied only
    to the exact shape `MARKDOWN_HEADING_RE` matches -- `#5 no space` is not
    that shape and is left alone. Table cells need none: a table block is split
    on the default separators and cited by its marker, so nothing reads a
    heading out of a row.
    """
    return MARKDOWN_HEADING_RE.sub(lambda match: "\\" + match.group(0), text)


def _one_line(text: str) -> str:
    """Collapse a run of whitespace -- a cell's own line breaks included -- to
    single spaces, so a table row is always exactly one line.

    The tabular chunker (`core/knowledge_base.py::_chunk_table_block`) reads
    the line after the marker as the column header and every later line as a
    row; a cell containing a newline would silently become an extra, shorter
    row. Applied to headings too, where a line break would otherwise split one
    heading into a heading plus an orphan line.
    """
    return " ".join(text.split())


def _parse_xml_bytes(data: bytes, name: str) -> str:
    try:
        # A separate pass collects the document's own namespace declarations;
        # the element tree itself only carries Clark-notation `{uri}local`
        # names, so the prefixes have to come from these start-ns events.
        ns_prefixes = {
            uri: prefix
            for _, (prefix, uri) in ET.iterparse(
                io.BytesIO(data), events=("start-ns",)
            )
        }
        root = ET.fromstring(data)
    except ET.ParseError as exc:
        raise ConfigurationError(
            f"Failed to parse XML file '{name}': {exc}"
        ) from exc

    lines = [f"[XML: {name}]"]
    _render_xml_tree(root, lines, ns_prefixes)
    return "\n".join(lines)


def _qualified_name(name: str, ns_prefixes: dict) -> str:
    """Render a Clark-notation `{uri}local` name as `prefix:local` using the
    document's own namespace declarations, falling back to the bare local
    name when no prefix is known (e.g. the default namespace)."""
    if not name.startswith("{"):
        return name
    uri, _, local = name[1:].partition("}")
    prefix = ns_prefixes.get(uri)
    return f"{prefix}:{local}" if prefix else local


def _normalize_xml_text(text: "str | None") -> str:
    return " ".join(text.split()) if text else ""


def _render_element_open_line(elem, depth: int, ns_prefixes: dict) -> str:
    indent = "  " * depth
    tag = _qualified_name(elem.tag, ns_prefixes)
    attrs = " ".join(
        f"{_qualified_name(k, ns_prefixes)}={quoteattr(v)}"
        for k, v in elem.attrib.items()
    )
    line = f"{indent}<{tag}" + (f" {attrs}" if attrs else "") + ">"
    text = _normalize_xml_text(elem.text)
    if text:
        line += f" {escape(text)}"
    return line


def _render_xml_tree(root, lines: list, ns_prefixes: dict) -> None:
    """Depth-first render of the element tree, including each element's tail
    text (mixed content following a child's closing tag). Iterative rather
    than recursive so a pathologically deep-but-narrow document can't blow
    the Python call stack -- an explicit stack is bounded only by memory,
    matching what `ET.parse` itself can already handle."""
    RENDER, TAIL = 0, 1
    stack = [(RENDER, root, 0)]
    while stack:
        kind, payload, depth = stack.pop()
        if kind == TAIL:
            lines.append(f"{'  ' * depth}{payload}")
            continue

        elem = payload
        lines.append(_render_element_open_line(elem, depth, ns_prefixes))

        items = []
        for child in elem:
            items.append((RENDER, child, depth + 1))
            tail = _normalize_xml_text(child.tail)
            if tail:
                items.append((TAIL, escape(tail), depth + 1))
        stack.extend(reversed(items))


# `csv.reader` refuses a field longer than 131,072 characters by default --
# a guard against unbounded memory growth while streaming a file, not a limit
# this application chose. Nothing here streams: the whole document is already
# in memory as a string before the reader sees it, and the upload routes admit
# a 30MB file (`ui/backend/knowledge_bases.py`), so the default rejected valid
# documents -- a CSV with one long notes column -- as unreadable. Raised once,
# at import, rather than set and restored around each parse: the setting is
# process-global and `ui/backend/ingestion.py` parses on a thread pool, where
# one document's restore would shrink the limit under another's parse.
_CSV_FIELD_LIMIT = 32 * 1024 * 1024
csv.field_size_limit(_CSV_FIELD_LIMIT)


def _parse_csv_bytes(data: bytes, name: str, *, lenient_text: bool = False) -> str:
    """Render a CSV as one marker line and one row per line.

    A CSV used to be decoded and handed on as prose, which cost it everything
    the same table gets as a `.xlsx`: the knowledge base repeats a table
    block's marker and column header at the top of every chunk it cuts, and
    without a marker a CSV was packed on the generic separators instead --
    losing the column names after the first chunk and cutting rows in half.
    The marker is what buys that back (`core/knowledge_base.py`).

    Read through `csv.reader` rather than split on newlines, because the two
    disagree exactly where it matters: Excel writes a cell containing a line
    break as a quoted field spanning two physical lines, and every column
    after such a break would otherwise sit under the wrong header. Written
    back through `csv.writer` for the same reason in reverse -- a value
    containing a comma has to stay one field, or it silently becomes two
    columns that no longer line up with the header row being repeated above
    it. Cell text is collapsed to a single line, so a row is always one line
    and stays one row all the way through chunking.

    Not sniffed: a semicolon- or tab-delimited export reads as one field per
    row and is rendered back unchanged, which is exactly what it did before.

    A row that renders as the marker's own shape is escaped, for the reason
    `_escape_marker_shaped` gives.
    """
    text = _decode_text(data, name, lenient=lenient_text)
    out = io.StringIO()
    writer = csv.writer(out, lineterminator="\n")
    try:
        for row in csv.reader(io.StringIO(text)):
            writer.writerow([_one_line(field) for field in row])
    except csv.Error as exc:
        # `ui/backend/ingestion.py` shows a failed document's message to the
        # customer who uploaded it, so it names the file rather than a byte
        # offset. It no longer guesses at a cause: an unbalanced quotation
        # mark does not come through here, the reader swallows the rest of
        # the file as one field, which is what it always did for a file under
        # the old limit. What is left is input the reader cannot get past at
        # all, an embedded NUL being the realistic one.
        raise ConfigurationError(f"Could not read '{name}' as CSV: {exc}.") from exc
    return f"[CSV: {name}]\n" + _escape_marker_shaped(out.getvalue().strip("\n"))


# Both caps exist because ingestion runs every customer's uploads on the one
# shared instance, so one workbook must not be able to buy minutes of CPU or
# gigabytes of RAM with a 30MB upload. Deliberately constants, not settings,
# for the same reason the PBKDF2 iteration count is: a limit that can be
# raised per-deployment will be, and then the instance it protects isn't
# protected. An .xlsx is a zip, so the archive's declared unpacked sizes are
# readable before any XML is parsed -- that is what catches a decompression
# bomb. The cell budget is the other axis: a sheet with one stray-formatted
# cell at row 1,048,576 declares every empty row above it, and `iter_rows`
# dutifully yields them all.
_MAX_XLSX_UNPACKED_BYTES = 300 * 1024 * 1024
_MAX_XLSX_CELLS = 5_000_000


def _parse_excel_bytes(data: bytes, name: str) -> str:
    try:
        import openpyxl
    except ImportError as exc:
        raise ConfigurationError(
            "Parsing Excel files requires the 'openpyxl' package. "
            "Install it with: pip install 'bestteam[tools-files]'"
        ) from exc

    try:
        with zipfile.ZipFile(io.BytesIO(data)) as archive:
            unpacked = sum(info.file_size for info in archive.infolist())
    except zipfile.BadZipFile as exc:
        raise ConfigurationError(
            f"Could not read '{name}' as an Excel workbook. "
            "Re-save it as .xlsx and upload it again."
        ) from exc
    if unpacked > _MAX_XLSX_UNPACKED_BYTES:
        raise ConfigurationError(
            f"'{name}' unpacks to more than "
            f"{_MAX_XLSX_UNPACKED_BYTES // (1024 * 1024)} MB of sheet data and "
            "was not ingested. Split the workbook into smaller files, or save "
            "the sheets you need as CSV."
        )

    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        parts = []
        cells_seen = 0
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Written back through `csv.writer` for the reason `_parse_csv_bytes`
            # gives: a cell containing a comma has to stay one field, or it
            # silently becomes two columns that no longer line up with the header
            # row repeated above it. `_one_line` keeps a cell containing a line
            # break one row for the same reason in the other axis.
            out = io.StringIO()
            writer = csv.writer(out, lineterminator="\n")
            for row in ws.iter_rows(values_only=True):
                cells_seen += len(row)
                if cells_seen > _MAX_XLSX_CELLS:
                    raise ConfigurationError(
                        f"'{name}' declares more than {_MAX_XLSX_CELLS:,} cells "
                        "and was not ingested. Delete unused rows and columns "
                        "(Excel keeps formatted-but-empty ones), or split the "
                        "workbook into smaller files."
                    )
                writer.writerow(["" if v is None else _one_line(str(v)) for v in row])
            parts.append(f"[Sheet: {sheet_name}]\n" + _escape_marker_shaped(out.getvalue().strip("\n")))
    finally:
        wb.close()
    return f"[Excel: {name}]\n\n" + "\n\n".join(parts)
